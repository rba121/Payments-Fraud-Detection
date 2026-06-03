import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime

import os
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "data", "payments.db")
OUTPUT_PATH = os.path.join(BASE_DIR, "reports", "Payment_Fraud_Report.xlsx")

def run_query(conn, sql):
    return pd.read_sql_query(sql, conn)

def style_header_row(ws, row, bg_color="1F4E79", font_color="FFFFFF"):
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color=font_color, name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def add_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="CCCCCC")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def auto_width(ws, padding=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + padding, 40)

def write_df_to_sheet(ws, df, start_row=1, flag_col=None, alt_row_color="EBF3FB"):
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row)

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Alternating row color
            if (r_idx - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", start_color=alt_row_color)
            # Highlight flagged rows red
            if flag_col and "flag_reason" in headers:
                flag_idx = headers.index("flag_reason")
                if row[flag_idx] and str(row[flag_idx]).strip():
                    cell.fill = PatternFill("solid", start_color="FFD7D7")
                    cell.font = Font(name="Arial", size=9, color="C00000")

    add_border(ws, start_row, start_row + len(df), 1, len(headers))
    auto_width(ws)

def build_report():
    conn = sqlite3.connect(DB_PATH)

    # --- Run all queries ---
    summary = run_query(conn, """
        SELECT COUNT(*) AS total_transactions,
               ROUND(SUM(amount),2) AS total_disbursed,
               SUM(is_flagged) AS flagged_count,
               ROUND(100.0*SUM(is_flagged)/COUNT(*),2) AS flag_rate_pct,
               ROUND(SUM(CASE WHEN is_flagged=1 THEN amount ELSE 0 END),2) AS flagged_amount
        FROM transactions
    """)

    fraud_breakdown = run_query(conn, """
        SELECT flag_reason, COUNT(*) AS occurrences,
               ROUND(SUM(amount),2) AS total_amount,
               ROUND(AVG(amount),2) AS avg_amount
        FROM transactions WHERE is_flagged=1
        GROUP BY flag_reason ORDER BY total_amount DESC
    """)

    overpayments = run_query(conn, """
        SELECT transaction_id, claimant_id, claimant_name, payment_type,
               amount, payment_date, region
        FROM transactions WHERE amount > 4500 ORDER BY amount DESC
    """)

    high_risk = run_query(conn, """
        SELECT claimant_id, claimant_name, COUNT(*) AS total_transactions,
               SUM(is_flagged) AS total_flags,
               ROUND(SUM(amount),2) AS total_paid,
               ROUND(SUM(CASE WHEN is_flagged=1 THEN amount ELSE 0 END),2) AS flagged_amount
        FROM transactions GROUP BY claimant_id, claimant_name
        HAVING SUM(is_flagged) >= 2 ORDER BY total_flags DESC
    """)

    regional = run_query(conn, """
        SELECT region, COUNT(*) AS total_payments,
               ROUND(SUM(amount),2) AS total_amount,
               SUM(is_flagged) AS flagged_count,
               ROUND(100.0*SUM(is_flagged)/COUNT(*),2) AS flag_rate_pct
        FROM transactions GROUP BY region ORDER BY flag_rate_pct DESC
    """)

    monthly = run_query(conn, """
        SELECT SUBSTR(payment_date,1,7) AS month,
               COUNT(*) AS total_payments,
               ROUND(SUM(amount),2) AS total_disbursed,
               SUM(is_flagged) AS flags
        FROM transactions GROUP BY month ORDER BY month
    """)

    payment_type = run_query(conn, """
        SELECT payment_type, COUNT(*) AS total, SUM(is_flagged) AS flagged,
               ROUND(100.0*SUM(is_flagged)/COUNT(*),2) AS fraud_rate_pct,
               ROUND(AVG(amount),2) AS avg_payment
        FROM transactions GROUP BY payment_type ORDER BY fraud_rate_pct DESC
    """)

    flagged_all = run_query(conn, """
        SELECT transaction_id, claimant_id, claimant_name, payment_type,
               amount, payment_date, region, flag_reason, processed_by
        FROM transactions WHERE is_flagged=1 ORDER BY payment_date DESC
    """)

    conn.close()

    # --- Build Workbook ---
    wb = openpyxl.Workbook()

    # ===================== SHEET 1: Executive Summary =====================
    ws1 = wb.active
    ws1.title = "Executive Summary"

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = "Payment Fraud Detection Report — WorkSafeBC"
    title_cell.font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:F2")
    sub = ws1["A2"]
    sub.value = f"Generated: {datetime.today().strftime('%B %d, %Y')}   |   Data Period: Jan 2024 – Dec 2024"
    sub.font = Font(italic=True, size=9, name="Arial", color="666666")
    sub.alignment = Alignment(horizontal="center")

    # KPI boxes
    ws1["A4"] = "Metric"
    ws1["B4"] = "Value"
    style_header_row(ws1, 4)

    kpi_labels = [
        ("Total Transactions Reviewed", summary["total_transactions"][0]),
        ("Total Amount Disbursed ($)", f"${summary['total_disbursed'][0]:,.2f}"),
        ("Flagged Transactions", summary["flagged_count"][0]),
        ("Flag Rate (%)", f"{summary['flag_rate_pct'][0]}%"),
        ("Total Flagged Amount ($)", f"${summary['flagged_amount'][0]:,.2f}"),
    ]

    highlight_colors = ["FFFFFF", "EBF3FB", "FFFFFF", "EBF3FB", "FFD7D7"]
    for i, (label, value) in enumerate(kpi_labels, start=5):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = value
        for col in ["A", "B"]:
            cell = ws1[f"{col}{i}"]
            cell.font = Font(name="Arial", size=10, bold=(col == "A"))
            cell.fill = PatternFill("solid", start_color=highlight_colors[i - 5])
            cell.alignment = Alignment(horizontal="left" if col == "A" else "right")

    add_border(ws1, 4, 9, 1, 2)
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22

    # Fraud type breakdown table
    ws1["A12"] = "Fraud Type Breakdown"
    ws1["A12"].font = Font(bold=True, size=11, name="Arial", color="1F4E79")

    write_df_to_sheet(ws1, fraud_breakdown, start_row=13)

    # ===================== SHEET 2: Overpayments =====================
    ws2 = wb.create_sheet("Overpayments")
    ws2["A1"] = "Transactions Exceeding Payment Threshold ($4,500)"
    ws2["A1"].font = Font(bold=True, size=12, name="Arial", color="C00000")
    ws2["A1"].alignment = Alignment(horizontal="left")
    write_df_to_sheet(ws2, overpayments, start_row=3)

    # ===================== SHEET 3: High Risk Claimants =====================
    ws3 = wb.create_sheet("High Risk Claimants")
    ws3["A1"] = "Claimants with 2+ Flagged Transactions"
    ws3["A1"].font = Font(bold=True, size=12, name="Arial", color="C00000")
    write_df_to_sheet(ws3, high_risk, start_row=3)

    # ===================== SHEET 4: Regional Analysis =====================
    ws4 = wb.create_sheet("Regional Analysis")
    ws4["A1"] = "Payment Fraud Rate by Region"
    ws4["A1"].font = Font(bold=True, size=12, name="Arial", color="1F4E79")
    write_df_to_sheet(ws4, regional, start_row=3)

    # Bar chart for regional fraud rate
    chart = BarChart()
    chart.type = "col"
    chart.title = "Flag Rate by Region (%)"
    chart.y_axis.title = "Flag Rate (%)"
    chart.x_axis.title = "Region"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws4, min_col=5, min_row=3, max_row=3 + len(regional))
    cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + len(regional))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws4.add_chart(chart, "H3")

    # ===================== SHEET 5: Monthly Trend =====================
    ws5 = wb.create_sheet("Monthly Trend")
    ws5["A1"] = "Monthly Payment Volume & Fraud Flags"
    ws5["A1"].font = Font(bold=True, size=12, name="Arial", color="1F4E79")
    write_df_to_sheet(ws5, monthly, start_row=3)

    # ===================== SHEET 6: Payment Type Analysis =====================
    ws6 = wb.create_sheet("Payment Type Analysis")
    ws6["A1"] = "Fraud Rate by Payment Type"
    ws6["A1"].font = Font(bold=True, size=12, name="Arial", color="1F4E79")
    write_df_to_sheet(ws6, payment_type, start_row=3)

    # ===================== SHEET 7: All Flagged Transactions =====================
    ws7 = wb.create_sheet("Flagged Transactions")
    ws7["A1"] = "All Flagged Transactions — Full Review Log"
    ws7["A1"].font = Font(bold=True, size=12, name="Arial", color="C00000")
    write_df_to_sheet(ws7, flagged_all, start_row=3, flag_col="flag_reason")

    wb.save(OUTPUT_PATH)
    print(f"Report saved to {OUTPUT_PATH}")
    print(f"  Total transactions: {summary['total_transactions'][0]}")
    print(f"  Flagged: {summary['flagged_count'][0]} ({summary['flag_rate_pct'][0]}%)")
    print(f"  Flagged amount: ${summary['flagged_amount'][0]:,.2f}")

if __name__ == "__main__":
    build_report()
